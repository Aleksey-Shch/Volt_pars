#import json
import requests
from bs4 import BeautifulSoup
import openpyxl


## чтение данных с сайта и сохранение в файле
# def reader_url_saved_text(url):
#     try:
#         headers = {
#             "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
#             "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.99 Safari/537.36"
#         }
#
#         reg = requests.get(url, headers=headers)
#         text_html = reg.text
#     except Exception:
#         print(f"Ошибка при чтении страницы {url}")
#     try:
#         with open("index.html", "w", encoding="utf-8") as file:
#             file.write(text_html)
#     except Exception:
#         print(f"Ошибка при сохранении файла")
#
# reader_url_saved_text("https://voltag.ru/catalog/group/voltag_alb0829_generator/")


#Запись в файл

# Чтение из файла и преобразовываем с soup
def read_text():
    with open('index.html', 'r', encoding='utf-8') as w_file:
        src = w_file.read()
    return BeautifulSoup(src, "lxml")

#Чтение данных из сохраненного файла
soup = read_text()

# фильтрация данных характеристики div class='catalog_group_components_c'>
quotes_components = soup.find_all('div', class_='catalog_group_components_c')#.find_all('tr')
#rint(quotes_components)
n=1
for i in quotes_components:
    href_comp = i.find('a').get('href')
    name_comp = i.find('a').text
    nomer_comp = i.find('div', class_='catalog_group_components_c_d').text
    print(f'{n} - href={href_comp},| имя={name_comp},| номер={nomer_comp}')
    n+=1
#https://voltag.ru/components/list/?q=ALB0829
#href = /catalog/group/voltag_ahb2132_schetki_generatora/, имя=AHB2132, номер = Щетки генератора
#https://voltag.ru/catalog/group/voltag_aeb5546_diodnij_most_generatora/?q=ALB0829
#https://voltag.ru/components/list/?q=ALA0236

# # запись данных в файл
# excel_file = openpyxl.load_workbook('team.xlsx')
# shet_names = excel_file.sheetnames
#
# if model in shet_names:
#     print('Yes')
#     excel_sheet = excel_file[model]
# #    excel_file.worksheets[excel_sheet].clear()
# else:
#     print('No')
#     excel_sheet = excel_file.create_sheet(title=model)
#
# #Установки ширины столбцов
# excel_sheet.column_dimensions["A"].width = 18
# excel_sheet.column_dimensions["B"].width = 18
# excel_sheet.column_dimensions["C"].width = 5
# excel_sheet.column_dimensions["D"].width = 5
# excel_sheet.column_dimensions["E"].width = 30
# excel_sheet.column_dimensions["F"].width = 8
# excel_sheet.column_dimensions["G"].width = 5
# excel_sheet.column_dimensions["H"].width = 5
# excel_sheet.column_dimensions["I"].width = 50
# # Запись сроссов в файл
#
# excel_sheet.cell(row=1, column=1).value = 'Аналоги -' # Шиниа 20 -добавить пожзже форматирование
# excel_sheet.cell(row=1, column=2).value = model
# i=3
# for katalo, nomra in cross.items():
#     if isinstance(nomra,str):
#         excel_sheet.cell(row=i, column=1).value = katalo
#         excel_sheet.cell(row=i, column=2).value = nomra
#         i+=1
#     else:
#         for nomer in nomra:
#             excel_sheet.cell(row=i, column=1).value = katalo
#             excel_sheet.cell(row=i, column=2).value = nomer
#             i += 1
#
# # """
# # Запись даннаых характеристики в файл
# excel_sheet.cell(row=1, column=4).value = 'Характеристика детали'
#
# i=3
# for harakter in haratkeristika:
#     r=4
#     for znacgenie in harakter:
#
#         excel_sheet.cell(row=i, column=r).value = znacgenie
#         r+=1
#     i+=1
#
# # Запись примениомсти
# excel_sheet.cell(row=1, column=9).value = 'Список применимости:'
#
# i=3
# for avto in primenimost:
#     excel_sheet.cell(row=i, column=9).value = ''.join(avto).strip()
#     i+=1
#
# excel_file.save('team.xlsx')