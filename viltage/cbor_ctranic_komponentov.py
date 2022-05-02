from bs4 import BeautifulSoup
from main import sup_save
import openpyxl
import os.path

def open_file_list_component(model):
    try:
        if os.path.isfile(f'{model}.xlsx'):  # Если файл сужествует открываем для записи
            excel_file = openpyxl.load_workbook(f'{model}.xlsx')
            shet_names = excel_file.sheetnames
            if (f'{model}_компонент') in shet_names:  # проверияем существует ли лист с компонентами
                print(f'Есть лист компонентов')
                #excel_sheet = excel_file.active
                excel_sheet = excel_file[(f'{model}_компонент')]
            else:
                print(f'Нету листа компонентов')
        else:  # Иначе ничего не открываем
            print(f'Нет такого файла')
        # Запись даннаых характеристики в файл
#        excel_sheet.cell(row=1, column=2).value = 'Список компонентов'
        print(excel_sheet.max_row)
        for i in range(excel_sheet.max_row-2):
            url_components = excel_sheet.cell(row=i+3, column=3).value
            sup_save(url_components)
        print('Все!')
    except Exception as error:
        print('Ошибка в формировании и сохранении файла: ' + repr(error))

if __name__ == '__main__':
    open_file_list_component("ALA2610")