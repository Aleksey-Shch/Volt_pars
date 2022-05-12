from bs4 import BeautifulSoup
import os.path
import requests
import re
import openpyxl

# чтение данных с сайта и формирыем сразу суп
def reader_url_component(url):
    try:
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Connection": "keep-alive",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36"
# другой комп           "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.4844.84 Safari/537.36 OPR/85.0.4341.75 (Edition Yx)"
        }
        reg = requests.get(url, headers=headers)
        text_html = reg.text
        return BeautifulSoup(text_html, "lxml")
    except Exception:
        print(f"Ошибка при чтении страницы {url}")


# поиск всех компонентов на странице с их ссылками
def saved_component_ctranica(soup):
    # проверка на наличие компонентов
    spisok_componentov={}
    if soup.find('div', class_ = 'catalog_item_title_wrap'):
        href_componenta = soup.find('div', class_='catalog_list').find_all('div', {'class': re.compile("catalog_item ")})
        #n = 1 catalog_item_title_wrap
        for s in href_componenta:  # [0]:
            #print(f'{n}' - {s}')
            #    href_comp = s.find('div', class_='catalog_item_title').get('href')
            href_comp = 'https://voltag.ru' + s.find('a').get('href')
            nomer_comp = (s.find('div', class_='catalog_item_title_wrap').text).replace("\n","")
            nazvanie_comp = s.find('div', class_='catalog_item_subtitle').text
            #print(f'{nomer_comp} - {nazvanie_comp} - {href_comp}') #, sep="\n")
            spisok_componentov[nomer_comp] = [nazvanie_comp, href_comp]
            # #Запись данных в файл формата json
            # with open(f'cross_{quotes_model}.json', 'w') as j_file:
            #     json.dump(cross, j_file, indent=4, ensure_ascii=False)
        return spisok_componentov
    print('Компонентов нет')

# Проверияем сколько страниц, если не одна перебираем
# все страницы и сохраняем список компонентов
def perebor_pages_component(soup):
    res, res_temp = {}, {}
    if soup.find('div', class_='page_number_outer'):
        print(f"Много листов ")
        res_temp=res.copy()
        res = {**res_temp, **(saved_component_ctranica(soup))}
        for s in soup.find('div', {'id': 'page_navigation'}).find_all('a'):
            # print(f"Много листов - {stranica}")
            soup_list = reader_url_component(f'https://voltag.ru{s.get("href")}')
            # тут надо сделать чтение всей станицы в файл
            res_temp = res.copy()
            res = {**res_temp, **(saved_component_ctranica(soup_list))}
    else:
        # тут надо сделать чтение всей станицы в файл
        # print(f"Один листов - {stranica}")
        res_temp = res.copy()
        res = {**res_temp, **(saved_component_ctranica(soup))}
    return res

def save_components(model, spisok):  #, haratkeristika, cross, primenimost):
    try:
        if os.path.isfile(f'{model}.xlsx'):  # Если файл сужествует открываем для записи
            excel_file = openpyxl.load_workbook(f'{model}.xlsx')
            shet_names = excel_file.sheetnames
            if (f'{model}_компонент') in shet_names:  # проверияем существует ли лист с такой деталью
                print(f'Есть такой лист. Сохранено как {model}new')
                excel_sheet = excel_file.create_sheet(title=(f'{model}_компонент1'))
                #excel_sheet = excel_file[model]
            else:
                #print('No')
                excel_sheet = excel_file.create_sheet(title=(f'{model}_компонент'))
        else:  # Иначе открываем пустой и формуем лист
            excel_file = openpyxl.Workbook()
            excel_sheet = excel_file.active
            excel_sheet.title = (f'{model}_компонент')
            #excel_sheet = excel_file.create_sheet(title=model) # новая страница, имя model
        # Запись даннаых характеристики в файл
        excel_sheet.cell(row=1, column=2).value = 'Список компонентов'

        stroka = 3
        for keys, values in spisok.items():
            #print(f'{keys} - {values}')
            excel_sheet.cell(row=stroka, column=1).value = keys
            excel_sheet.cell(row=stroka, column=2).value = values[0]
            excel_sheet.cell(row=stroka, column=3).value = values[1]
            stroka += 1
        excel_file.save(f'{model}.xlsx')

    except Exception as error:
        print('Ошибка в формировании и сохранении файла: ' + repr(error))

if __name__ == '__main__':
    spisok_componentov={}
    model = 'ALA2610' # три листа компонентовa
    #model = 'ala3231' 'ala2610' 'ALA0785' 'ALA0879' # нет компонентов совсем
    #Сохраняем первую страницу
    soup = reader_url_component(f"https://voltag.ru/components/list/p-1/?q={model}")
    perebor_pages_component(soup)
    save_components(model, spisok_componentov)

    for keys, values in spisok_componentov.items():
        print(f'{keys} - {values}')
