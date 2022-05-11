import json
import os.path
import requests
from bs4 import BeautifulSoup
import openpyxl
import componenty_detali


# чтение данных с сайта и сохранение в файле index.html
def reader_url_saved_text(url):
    try:
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
            "Accept-Encoding": "gzip, deflate, br",
            "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
            "Connection": "keep-alive",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.82 Safari/537.36"
        }
        reg = requests.get(url, headers=headers)
        text_html = reg.text
    except Exception:
        print(f"Ошибка при чтении страницы {url}")
    try:
        with open("index.html", "w", encoding="utf-8") as file:
            file.write(text_html)
    except Exception:
        print(f"Ошибка при сохранении файла")

# Чтение из файла и преобразовываем с soup return BeautifulSoup
def read_text():
    try:
        with open("index.html", "r", encoding="utf-8") as w_file:
            src = w_file.read()
        return BeautifulSoup(src, "lxml")
    except Exception as err:
        print(f"Ошибка чтения из файла. Ошибка {err}")

def in_components(soup):
    if (soup.find('span', class_='ninf-cl').text) == "все":
        return "Есть"

# фильтрация названия модели
def filter_model(soup):
    model = ''.join(soup.find('div', class_='catalog_group_title').text.split())
    return model

# фильтрация данных характеристики
def filter_harakteristika(soup):
    quotes_harakteristika = soup.find('div', class_='catalog_group_params').find_all('tr')
    # форматирование данных характеристики haratkeristika
    haratkeristika = []
    for znac in quotes_harakteristika:
        data = [x.get_text().replace('\xa0', ' ') for x in znac.find_all('td')]
        if len(data) != 0:
            haratkeristika.append(data)
    return haratkeristika

# фильтрация данных кроссов
def filter_kross(soup):
    quotes_kross = soup.find('div', class_='catalog_group_crosslist_info')

    # форматирование данных кроссов cross
    cross = {}
    katal = 'Zero'
    for td in quotes_kross.find_all('td'):
        znac = td.text.strip()
        if str(td).find('mnfr') == -1:
            if len(znac.split(', ')) == 1:
                cross[katal] = znac
            else:
                for td1 in (znac.split(', ')):
                    cross.setdefault(katal, []).append(td1)
        else:
            if znac == '':
                katal = 'Zero'
            else:
                katal = znac
    return cross

# фильтраци данных применимости
def filter_primenomost(soup):
    quotes_primenimost = soup.find('div', class_='catalog_group_application_info')
    # форматирование данных применимости primenimost
    primenimost = [[x.replace('\t', ' ')] for x in quotes_primenimost.text.split('\n')]
    return primenimost

# запись данных в файл
def save_dannix_detali(model, haratkeristika, cross, primenimost, list='model'):

    try:
        if os.path.isfile(f'{model}.xlsx'):  # Если файл сужествует открываем для записи
            excel_file = openpyxl.load_workbook(f'{model}.xlsx')
            shet_names = excel_file.sheetnames
            if list in shet_names:  # проверияем существует ли лист с такой деталью
                print(f'Есть такой лист. Сохранено как {list}new')
                excel_sheet = excel_file.create_sheet(title=(f'{list}NEW'))
                #excel_sheet = excel_file[model]
            else:
                #print('No')
                excel_sheet = excel_file.create_sheet(title=list)
        else:  # Иначе открываем пустой и формуем лист
            excel_file = openpyxl.Workbook()
            excel_sheet = excel_file.active
            excel_sheet.title = list
            #excel_sheet = excel_file.create_sheet(title=model) # новая страница, имя model

        # Установки ширины столбцов
        excel_sheet.column_dimensions["A"].width = 5
        excel_sheet.column_dimensions["B"].width = 30
        excel_sheet.column_dimensions["C"].width = 12

        # Запись даннаых характеристики в файл
        excel_sheet.cell(row=1, column=2).value = 'Характеристика детали'

        stroka = 3
        for harakter in haratkeristika:
            stolb = 1
            for znacgenie in harakter:
                excel_sheet.cell(row=stroka, column=stolb).value = znacgenie
                stolb += 1
            stroka += 1

        # Запись сроссов в файл
        excel_sheet.cell(row=stroka + 1, column=2).value = 'Аналоги -'  # Шиниа 20 -добавить пожзже форматирование
        excel_sheet.cell(row=stroka + 1, column=3).value = model
        stroka += 3
        for katalo, nomra in cross.items():
            if isinstance(nomra, str):
                excel_sheet.cell(row=stroka, column=2).value = katalo
                excel_sheet.cell(row=stroka, column=3).value = nomra
                stroka += 1
            else:
                for nomer in nomra:
                    excel_sheet.cell(row=stroka, column=2).value = katalo
                    excel_sheet.cell(row=stroka, column=3).value = nomer
                    stroka += 1

        # Запись применимости
        excel_sheet.cell(row=stroka+2, column=2).value = 'Список применимости:'

        stroka += 3
        for avto in primenimost:
            excel_sheet.cell(row=stroka, column=2).value = ''.join(avto).strip()
            stroka += 1

        excel_file.save(f'{model}.xlsx')
    except Exception as error:
        print('Ошибка в формировании и сохранении файла: ' + repr(error))

def sup_save(url, model_osnova=None):
    if len(url) > 20 and url.find("https://voltag.ru") == 0:
        if model_osnova == None:
            reader_url_saved_text(url)  # сохраняем новую страницу
            soup = read_text()  # делаем суп
            save_dannix_detali(filter_model(soup), filter_harakteristika(soup), filter_kross(soup),
                               filter_primenomost(soup))
        else:
            reader_url_saved_text(url)  # сохраняем новую страницу
            soup = read_text()  # делаем суп
            save_dannix_detali(model_osnova, filter_harakteristika(soup), filter_kross(soup),
                               filter_primenomost(soup),filter_model(soup))
    else:
        soup = read_text()  # делаем суп
        save_dannix_detali(filter_model(soup), filter_harakteristika(soup), filter_kross(soup),
                           filter_primenomost(soup))
    return filter_model(soup)


# #Запись данных в файл формата json
# with open(f'cross_{quotes_model}.json', 'w') as j_file:
#     json.dump(cross, j_file, indent=4, ensure_ascii=False)

# reader_url_saved_text("https://voltag.ru/catalog/group/voltag_ala0236_generator/?q=ALA0236")
# reader_url_saved_text("https://voltag.ru/catalog/list/voltag_ala2610_generator/?q=ala2610")
#https://voltag.ru/catalog/group/voltag_ala0879_generator/ # 3 страницы компонентов
#https://voltag.ru/components/list/?q=ALA0879
#https://voltag.ru/components/list/p-2/?q=ALA0879
#https://voltag.ru/components/list/p-3/?q=ALA0879
# Вводим адрес и созраняем в фалйе через функцию

#in_components(soup)
if __name__ == '__main__':
#    url_detali = input(f"Введите адрес страница с сата voltag.ru или просто Enter \n")
#    url_detali = 'https://voltag.ru/catalog/group/voltag_ala0879_generator/?q=ala0879'
#    url_detali = 'https://voltag.ru/catalog/list/voltag_ala2610_generator/?q=ala2610'
    url_detali = 'https://voltag.ru/catalog/group/voltag_ala0236_generator/?q=ALA0236'
    model_osnovnaya = sup_save(url_detali)
#    sup_save("https://voltag.ru/catalog/group/voltag_ala0879_generator/?q=ala0879")
    soup = componenty_detali.reader_url_component(f"https://voltag.ru/components/list/p-1/?q={model_osnovnaya}")
    spisok_componentov_full = componenty_detali.perebor_pages_component(soup)
    componenty_detali.save_components(model_osnovnaya , spisok_componentov_full)
    for keys, values in spisok_componentov_full.items():
        model_comp = sup_save(values[1], model_osnovnaya)
    print('Готово')